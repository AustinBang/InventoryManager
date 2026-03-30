import os
import hashlib
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook

FILE_PATH = "rims.xlsx"

SHEETS = {
    "Users": ["username", "password_hash", "role"],
    "Products": ["sku", "name", "unit", "unit_cost", "on_hand", "low_stock_threshold"],
    "StockMovements": ["timestamp", "sku", "change_qty", "reason", "user"],
}


ROLES = {"Manager", "Chef", "Staff"}  # simple RBAC baseline


def hash_password(password: str) -> str:
    # Simple SHA-256 hash for a class project starter.
    # If this becomes “real”, switch to bcrypt/argon2.
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def ensure_workbook(path: str) -> None:
    if os.path.exists(path):
        return

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)

    # Create a default admin user
    ws_users = wb["Users"]
    ws_users.append(["admin", hash_password("admin123"), "Manager"])

    wb.save(path)


def load_wb():
    ensure_workbook(FILE_PATH)
    return load_workbook(FILE_PATH)


def find_row_by_value(ws, col_idx: int, value: str):
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx).value
        if str(cell).strip() == value:
            return r
    return None


class RIMSApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("RIMS - Restaurant Inventory (Excel Edition)")
        self.geometry("980x600")

        self.current_user = None
        self.current_role = None

        self.container = ttk.Frame(self)
        self.container.pack(fill="both", expand=True)

        self.show_login()

    def show_login(self):
        for w in self.container.winfo_children():
            w.destroy()
        LoginFrame(self.container, self).pack(fill="both", expand=True)

    def show_main(self):
        for w in self.container.winfo_children():
            w.destroy()
        MainFrame(self.container, self).pack(fill="both", expand=True)


class LoginFrame(ttk.Frame):
    def __init__(self, parent, app: RIMSApp):
        super().__init__(parent, padding=20)
        self.app = app

        ttk.Label(self, text="Login", font=("Segoe UI", 18, "bold")).pack(anchor="w")

        form = ttk.Frame(self)
        form.pack(pady=20, anchor="w")

        ttk.Label(form, text="Username").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Label(form, text="Password").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)

        self.username = ttk.Entry(form, width=30)
        self.password = ttk.Entry(form, width=30, show="*")
        self.username.grid(row=0, column=1, pady=6)
        self.password.grid(row=1, column=1, pady=6)

        btns = ttk.Frame(self)
        btns.pack(anchor="w", pady=10)

        ttk.Button(btns, text="Sign in", command=self.login).pack(side="left")

        hint = ttk.Label(
            self,
            text="Starter admin: admin / admin123 (change later)",
            foreground="#555",
        )
        hint.pack(anchor="w", pady=(10, 0))

    def login(self):
        u = self.username.get().strip()
        p = self.password.get().strip()
        if not u or not p:
            messagebox.showerror("Error", "Enter username + password.")
            return

        wb = load_wb()
        ws = wb["Users"]

        for r in range(2, ws.max_row + 1):
            username = str(ws.cell(r, 1).value).strip()
            pw_hash = str(ws.cell(r, 2).value).strip()
            role = str(ws.cell(r, 3).value).strip()

            if username == u and pw_hash == hash_password(p):
                self.app.current_user = u
                self.app.current_role = role
                self.app.show_main()
                return

        messagebox.showerror("Login failed", "Invalid username/password.")


class MainFrame(ttk.Frame):
    def __init__(self, parent, app: RIMSApp):
        super().__init__(parent, padding=12)
        self.app = app

        top = ttk.Frame(self)
        top.pack(fill="x")

        ttk.Label(
            top,
            text=f"RIMS (Excel) — User: {app.current_user}  Role: {app.current_role}",
            font=("Segoe UI", 12, "bold"),
        ).pack(side="left")

        ttk.Button(top, text="Logout", command=app.show_login).pack(side="right")

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, pady=(10, 0))

        self.products_tab = ProductsTab(self.notebook, app)
        self.movements_tab = MovementsTab(self.notebook, app)

        self.notebook.add(self.products_tab, text="Products / Stock")
        self.notebook.add(self.movements_tab, text="Stock Movements (Ledger)")


class ProductsTab(ttk.Frame):
    def __init__(self, parent, app: RIMSApp):
        super().__init__(parent, padding=10)
        self.app = app

        # Controls row
        ctrl = ttk.Frame(self)
        ctrl.pack(fill="x", pady=(0, 10))

        self.search_var = tk.StringVar()
        ttk.Label(ctrl, text="Search SKU/Name:").pack(side="left")
        ttk.Entry(ctrl, textvariable=self.search_var, width=30).pack(side="left", padx=8)
        ttk.Button(ctrl, text="Search", command=self.refresh).pack(side="left")

        ttk.Button(ctrl, text="Add Product", command=self.add_product_dialog).pack(side="right")

        # Treeview
        cols = ("sku", "name", "unit", "unit_cost", "on_hand", "low_stock_threshold", "status")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=120, anchor="w")
        self.tree.column("name", width=220)
        self.tree.column("status", width=120)

        self.tree.pack(fill="both", expand=True)

        # Buttons
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", pady=10)

        ttk.Button(bottom, text="Adjust Stock (Sale/Purchase/Waste)", command=self.adjust_stock_dialog).pack(side="left")
        ttk.Button(bottom, text="Refresh", command=self.refresh).pack(side="right")

        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        wb = load_wb()
        ws = wb["Products"]

        q = self.search_var.get().strip().lower()

        for r in range(2, ws.max_row + 1):
            sku = str(ws.cell(r, 1).value or "").strip()
            name = str(ws.cell(r, 2).value or "").strip()
            unit = str(ws.cell(r, 3).value or "").strip()
            unit_cost = float(ws.cell(r, 4).value or 0)
            on_hand = float(ws.cell(r, 5).value or 0)
            low_thr = float(ws.cell(r, 6).value or 0)

            if q and (q not in sku.lower() and q not in name.lower()):
                continue

            status = "LOW" if on_hand <= low_thr else "OK"
            iid = self.tree.insert("", "end", values=(sku, name, unit, unit_cost, on_hand, low_thr, status))
            if status == "LOW":
                self.tree.item(iid, tags=("low",))

        self.tree.tag_configure("low", background="#ffe7e7")

    def add_product_dialog(self):
        if self.app.current_role not in {"Manager", "Chef"}:
            messagebox.showerror("Not allowed", "Only Manager/Chef can add products.")
            return

        dlg = tk.Toplevel(self)
        dlg.title("Add Product")
        dlg.geometry("420x320")
        dlg.transient(self)
        dlg.grab_set()

        fields = {}
        form = ttk.Frame(dlg, padding=12)
        form.pack(fill="both", expand=True)

        def add_row(label, key):
            row = ttk.Frame(form)
            row.pack(fill="x", pady=6)
            ttk.Label(row, text=label, width=18).pack(side="left")
            e = ttk.Entry(row)
            e.pack(side="left", fill="x", expand=True)
            fields[key] = e

        add_row("SKU", "sku")
        add_row("Name", "name")
        add_row("Unit (e.g. lb)", "unit")
        add_row("Unit Cost", "unit_cost")
        add_row("On Hand", "on_hand")
        add_row("Low Stock Threshold", "low_stock_threshold")

        def save():
            sku = fields["sku"].get().strip()
            name = fields["name"].get().strip()
            unit = fields["unit"].get().strip()
            if not sku or not name:
                messagebox.showerror("Error", "SKU and Name are required.")
                return

            try:
                unit_cost = float(fields["unit_cost"].get().strip() or 0)
                on_hand = float(fields["on_hand"].get().strip() or 0)
                low_thr = float(fields["low_stock_threshold"].get().strip() or 0)
            except ValueError:
                messagebox.showerror("Error", "Cost/quantities must be numbers.")
                return

            wb = load_wb()
            ws = wb["Products"]

            if find_row_by_value(ws, 1, sku) is not None:
                messagebox.showerror("Error", f"SKU {sku} already exists.")
                return

            ws.append([sku, name, unit, unit_cost, on_hand, low_thr])
            wb.save(FILE_PATH)
            dlg.destroy()
            self.refresh()

        ttk.Button(form, text="Save", command=save).pack(anchor="e", pady=(12, 0))

    def adjust_stock_dialog(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Select a product", "Click a product row first.")
            return

        sku = self.tree.item(sel[0], "values")[0]

        dlg = tk.Toplevel(self)
        dlg.title(f"Adjust Stock — {sku}")
        dlg.geometry("420x240")
        dlg.transient(self)
        dlg.grab_set()

        form = ttk.Frame(dlg, padding=12)
        form.pack(fill="both", expand=True)

        ttk.Label(form, text=f"SKU: {sku}", font=("Segoe UI", 11, "bold")).pack(anchor="w")

        qty_var = tk.StringVar()
        reason_var = tk.StringVar(value="Sale")

        row1 = ttk.Frame(form)
        row1.pack(fill="x", pady=8)
        ttk.Label(row1, text="Change Qty (+/-)", width=18).pack(side="left")
        ttk.Entry(row1, textvariable=qty_var).pack(side="left", fill="x", expand=True)

        row2 = ttk.Frame(form)
        row2.pack(fill="x", pady=8)
        ttk.Label(row2, text="Reason", width=18).pack(side="left")
        reasons = ["Sale", "Purchase", "Waste", "Prep", "CountCorrection"]
        ttk.Combobox(row2, textvariable=reason_var, values=reasons, state="readonly").pack(side="left", fill="x", expand=True)

        def apply():
            try:
                change = float(qty_var.get().strip())
            except ValueError:
                messagebox.showerror("Error", "Quantity must be a number.")
                return

            wb = load_wb()
            ws_prod = wb["Products"]
            ws_mov = wb["StockMovements"]

            r = find_row_by_value(ws_prod, 1, sku)
            if r is None:
                messagebox.showerror("Error", "Product not found.")
                return

            on_hand = float(ws_prod.cell(r, 5).value or 0)
            new_on_hand = on_hand + change
            if new_on_hand < 0:
                messagebox.showerror("Error", "Stock cannot go below 0 (for this starter version).")
                return

            # Update product
            ws_prod.cell(r, 5).value = new_on_hand

            # Append to ledger
            ws_mov.append([
                datetime.now().isoformat(timespec="seconds"),
                sku,
                change,
                reason_var.get().strip(),
                self.app.current_user,
            ])

            wb.save(FILE_PATH)
            dlg.destroy()
            self.refresh()

        ttk.Button(form, text="Apply", command=apply).pack(anchor="e", pady=(12, 0))


class MovementsTab(ttk.Frame):
    def __init__(self, parent, app: RIMSApp):
        super().__init__(parent, padding=10)
        self.app = app

        top = ttk.Frame(self)
        top.pack(fill="x", pady=(0, 10))

        ttk.Button(top, text="Refresh", command=self.refresh).pack(side="right")

        cols = ("timestamp", "sku", "change_qty", "reason", "user")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=160, anchor="w")
        self.tree.column("timestamp", width=200)
        self.tree.pack(fill="both", expand=True)

        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        wb = load_wb()
        ws = wb["StockMovements"]

        # show latest first
        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append([
                ws.cell(r, 1).value,
                ws.cell(r, 2).value,
                ws.cell(r, 3).value,
                ws.cell(r, 4).value,
                ws.cell(r, 5).value,
            ])

        for data in reversed(rows[-500:]):  # cap for UI
            self.tree.insert("", "end", values=tuple(data))


if __name__ == "__main__":
    ensure_workbook(FILE_PATH)
    app = RIMSApp()
    app.mainloop()


